"""
excel_filler.py
업로드된 비용청구 .xlsm 의 '작성시트'에 영수증 데이터를 채운다.

★ 도형 보존 방식 ★
openpyxl 로 재저장하면 버튼 그림·도형(텍스트박스/그룹)·VML 등 drawing 객체가
사라진다. 그래서 이 모듈은 .xlsm 을 zip 으로 열어 해당 시트의 XML 에서
대상 셀 값만 직접 교체한다. 나머지(도형/이미지/VBA/서식/수식)는 원본 그대로
복사되므로 100% 보존된다.

채우는 열: C(영수일자), D(거래처명), E(목적), F(영수금액),
           G(결제방식), J(영수시간)
목적/결제방식은 양식의 드롭다운 목록에서 사용자가 고른 값을 그대로 기재한다.
"""

import re
import zipfile
from copy import copy  # noqa: F401  (호환용 import 유지)
from datetime import datetime, date, time
from io import BytesIO
from xml.sax.saxutils import escape

import openpyxl

SHEET_NAME = "작성시트"
CLAIM_SHEET = "비용청구서"   # 기초정보(소속/성명/카드/제목)·서명 상태가 들어가는 시트
SUPPORT_SHEET = "비용지원안내"  # 목적별 지원한도 안내 시트
FIRST_DATA_ROW = 15        # 실제 데이터 시작 행
EXCEL_EPOCH = date(1899, 12, 30)

# 비용청구서 시트 데이터 영역(매크로 '비용청구생성공통단계' 기준)
CLAIM_FIRST_ROW = 17       # 비용청구서 시트 데이터 시작 행
CLAIM_MAX_ROWS = 25        # 양식에 준비된 데이터행 수(17~41). 초과 시 행삽입 필요(미지원)

# 식대 카테고리: 매크로 '동작매크로_매크로검토' 3단계에서 '참여자 인원수 × 1인당 한도'로
# 검증하는 목적 목록(VBA 하드코딩과 동일). 1인당 한도 금액은 get_support_limits(현재값)에서 조회.
MEAL_PURPOSES = {
    "야근식비",
    "외근식비(조식)", "외근식비(중식)", "외근식비(석식)",
    "파견식비(조식)", "파견식비(중식.자택출퇴근)", "파견식비(중식.현지숙박)", "파견식비(석식)",
    "특근식비(조식)", "특근식비(중식)", "특근식비(석식)",
    "회식비(항목사용금지)",
    "주간스터디", "야간스터디",
    "인턴중식비", "인턴음료비",
}


# ---------------------------------------------------------------- 변환 함수
def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().replace(".", "-").replace("/", "-")
    s = "-".join(p for p in s.split("-") if p != "")
    for fmt in ("%Y-%m-%d", "%y-%m-%d", "%m-%d"):
        try:
            d = datetime.strptime(s, fmt)
            if fmt == "%m-%d":
                d = d.replace(year=datetime.now().year)
            return d.date()
        except ValueError:
            continue
    return None


def _to_time(v):
    if v is None or v == "":
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    s = str(v).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _to_amount(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    s = "".join(ch for ch in str(v) if ch.isdigit())
    return int(s) if s else None


def _date_serial(d: date) -> int:
    return (d - EXCEL_EPOCH).days


def _time_fraction(t: time) -> float:
    return (t.hour * 3600 + t.minute * 60 + t.second) / 86400.0


# ---------------------------------------------------------------- zip 헬퍼
def _read_bytes(src):
    if isinstance(src, (bytes, bytearray)):
        return bytes(src)
    if hasattr(src, "read"):
        return src.read()
    with open(src, "rb") as f:
        return f.read()


def _sheet_path_for(zf: zipfile.ZipFile, sheet_name: str) -> str:
    """workbook.xml + rels 를 읽어 시트 이름 -> worksheets/sheetN.xml 매핑."""
    wb = zf.read("xl/workbook.xml").decode("utf-8")
    rid = None
    for m in re.finditer(r'<sheet [^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wb):
        if m.group(1) == sheet_name:
            rid = m.group(2)
            break
    if rid is None:
        # name 과 r:id 순서가 바뀐 경우 대비
        for m in re.finditer(r'<sheet [^>]*r:id="([^"]+)"[^>]*name="([^"]+)"', wb):
            if m.group(2) == sheet_name:
                rid = m.group(1)
                break
    if rid is None:
        raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")

    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    for m in re.finditer(r'<Relationship [^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', rels):
        if m.group(1) == rid:
            target = m.group(2)
            return "xl/" + target.lstrip("/").replace("xl/", "", 1) if not target.startswith("xl/") else target
    # 일반적 형태: Target="worksheets/sheetN.xml"
    for m in re.finditer(r'<Relationship [^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', rels):
        if m.group(1) == rid:
            return "xl/" + m.group(2)
    raise ValueError(f"시트 경로를 찾을 수 없습니다: rid={rid}")


# ---------------------------------------------------------------- 셀 편집
def _cell_block(xml: str, coord: str):
    """coord(예: 'C25') 셀의 (start, end, style, has_value) 반환. 없으면 None."""
    m = re.search(r'<c r="%s"((?:\s+[^>]*?)?)(/>|>.*?</c>)' % re.escape(coord), xml, re.S)
    if not m:
        return None
    attrs = m.group(1)
    sm = re.search(r's="(\d+)"', attrs)
    style = sm.group(1) if sm else None
    has_value = "<v>" in m.group(2) or "<is>" in m.group(2)
    return m.start(), m.end(), style, has_value


def _build_cell(coord, style, kind, value):
    s_attr = f' s="{style}"' if style is not None else ""
    if kind == "num":
        return f'<c r="{coord}"{s_attr}><v>{value}</v></c>'
    if kind == "str":
        txt = escape(str(value))
        return (f'<c r="{coord}"{s_attr} t="inlineStr">'
                f'<is><t xml:space="preserve">{txt}</t></is></c>')
    raise ValueError(kind)


def _insert_cell(xml: str, coord: str, cell_xml: str) -> str:
    """XML 에 없는 셀을 해당 행의 올바른 열 위치(컬럼 순서)에 삽입한다.
    행이 없으면 행 자체를 행 번호 순서에 맞게 새로 만든다."""
    row = int(re.search(r"\d+", coord).group(0))
    col_idx = _col_to_idx(re.match(r"[A-Z]+", coord).group(0))

    rm = re.search(r'<row r="%d"(?:\s+[^>]*?)?>(.*?)</row>' % row, xml, re.S)
    if rm:
        inner = rm.group(1)
        pos = len(inner)  # 기본: 행 끝
        for cm in re.finditer(r'<c r="([A-Z]+)\d+"', inner):
            if _col_to_idx(cm.group(1)) > col_idx:
                pos = cm.start()
                break
        new_inner = inner[:pos] + cell_xml + inner[pos:]
        return xml[:rm.start(1)] + new_inner + xml[rm.end(1):]

    # self-closing <row r="N"/> 형태
    rm2 = re.search(r'<row r="%d"((?:\s+[^>]*?)?)/>' % row, xml)
    if rm2:
        replacement = f'<row r="{row}"{rm2.group(1)}>{cell_xml}</row>'
        return xml[:rm2.start()] + replacement + xml[rm2.end():]

    # 행 자체가 없으면 행 번호 순서를 지켜 새 행 삽입
    new_row = f'<row r="{row}">{cell_xml}</row>'
    for rm3 in re.finditer(r'<row r="(\d+)"', xml):
        if int(rm3.group(1)) > row:
            return xml[:rm3.start()] + new_row + xml[rm3.start():]
    return xml.replace("</sheetData>", new_row + "</sheetData>", 1)


def _set_cell(xml: str, coord: str, kind: str, value) -> str:
    blk = _cell_block(xml, coord)
    if blk is None:
        # 빈 셀이라 XML 에 없으면 새로 삽입
        return _insert_cell(xml, coord, _build_cell(coord, None, kind, value))
    start, end, style, _ = blk
    new = _build_cell(coord, style, kind, value)
    return xml[:start] + new + xml[end:]


def _force_full_recalc(workbook_xml: str) -> str:
    """xl/workbook.xml 의 <calcPr> 에 fullCalcOnLoad="1" 을 넣어,
    파일을 열 때 엑셀이 모든 수식을 강제로 재계산하게 한다.
    (값 셀만 직접 교체하면 수식 셀의 캐시값이 그대로 남아 재계산되지 않는다.)"""
    m = re.search(r"<calcPr\b[^>]*/>", workbook_xml)
    if m:
        tag = m.group(0)
        if "fullCalcOnLoad" in tag:
            tag = re.sub(r'fullCalcOnLoad="[^"]*"', 'fullCalcOnLoad="1"', tag)
        else:
            tag = tag[:-2] + ' fullCalcOnLoad="1"/>'
        return workbook_xml[:m.start()] + tag + workbook_xml[m.end():]
    # calcPr 가 없으면 sheets 뒤에 삽입 (calcPr 는 sheets 다음 위치).
    # <sheets>...</sheets> 와 자체닫힘 <sheets/> 둘 다 처리.
    return re.sub(r"(</sheets>|<sheets\b[^>]*/>)",
                  r'\1<calcPr fullCalcOnLoad="1"/>', workbook_xml, count=1)


def _first_empty_row(xml: str) -> int:
    r = FIRST_DATA_ROW
    while True:
        c = _cell_block(xml, f"C{r}")
        d = _cell_block(xml, f"D{r}")
        c_empty = (c is None) or (not c[3])
        d_empty = (d is None) or (not d[3])
        if c_empty and d_empty:
            return r
        r += 1


# ---------------------------------------------------------------- 드롭다운 목록
def _shared_strings(zf: zipfile.ZipFile):
    """sharedStrings.xml -> index 순서의 문자열 리스트."""
    try:
        xml = zf.read("xl/sharedStrings.xml").decode("utf-8")
    except KeyError:
        return []
    out = []
    for si in re.finditer(r"<si>(.*?)</si>", xml, re.S):
        body = si.group(1)
        # <si> 안의 모든 <t>...</t> 합치기 (rich text 대응)
        text = "".join(re.findall(r"<t[^>]*>(.*?)</t>", body, re.S))
        text = (text.replace("&amp;", "&").replace("&lt;", "<")
                .replace("&gt;", ">").replace("&quot;", '"').replace("&apos;", "'"))
        out.append(text)
    return out


def _col_to_idx(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch.upper()) - 64)
    return n


def _read_range_column(zf, sheet_name, col, row1, row2):
    """특정 시트의 한 열(col) row1~row2 셀 값을 리스트로 반환."""
    path = _sheet_path_for(zf, sheet_name)
    xml = zf.read(path).decode("utf-8")
    shared = _shared_strings(zf)
    out = []
    for r in range(row1, row2 + 1):
        coord = f"{col}{r}"
        m = re.search(r'<c r="%s"((?:\s+[^>]*?)?)(?:/>|>(.*?)</c>)' % re.escape(coord),
                      xml, re.S)
        if not m:
            continue
        attrs, inner = m.group(1), (m.group(2) or "")
        t = re.search(r't="([^"]+)"', attrs)
        ttype = t.group(1) if t else None
        vm = re.search(r"<v>(.*?)</v>", inner, re.S)
        if ttype == "s" and vm:
            idx = int(vm.group(1))
            if 0 <= idx < len(shared):
                out.append(shared[idx])
        elif ttype == "inlineStr":
            im = re.search(r"<t[^>]*>(.*?)</t>", inner, re.S)
            if im:
                out.append(im.group(1))
        elif vm:
            out.append(vm.group(1))
    return [v for v in (s.strip() for s in out) if v]


def get_dropdown_options(src_path_or_bytes):
    """
    양식에서 목적(E)·결제방식(G) 드롭다운 선택지를 추출한다.
    반환: {"purpose": [...], "payment": [...]}
    파싱 실패 시 빈 리스트.
    """
    raw = _read_bytes(src_path_or_bytes)
    zf = zipfile.ZipFile(BytesIO(raw))
    sheet_path = _sheet_path_for(zf, SHEET_NAME)
    xml = zf.read(sheet_path).decode("utf-8")

    payment, purpose = [], []

    # 결제방식(G): 표준 dataValidation list, formula1 이 따옴표 문자열
    for m in re.finditer(r"<dataValidation\b[^>]*?type=\"list\"[^>]*?>(.*?)</dataValidation>",
                         xml, re.S):
        block = m.group(0)
        sq = re.search(r'sqref="([^"]*)"', block)
        if sq and re.search(r'\bG\d', sq.group(1)):
            f1 = re.search(r"<formula1>\"?(.*?)\"?</formula1>", m.group(1), re.S)
            if f1:
                payment = [s.strip() for s in f1.group(1).split(",") if s.strip()]
            break

    # 목적(E): x14 dataValidation, formula1 이 다른 시트 범위 참조
    mx = re.search(r"<x14:dataValidation\b.*?</x14:dataValidation>", xml, re.S)
    for m in re.finditer(r"<x14:dataValidation\b(.*?)</x14:dataValidation>", xml, re.S):
        block = m.group(1)
        sq = re.search(r"<xm:sqref>([^<]*)</xm:sqref>", block)
        if sq and re.search(r'\bE\d', sq.group(1)):
            f = re.search(r"<xm:f>(.*?)</xm:f>", block, re.S)
            if f:
                ref = f.group(1)  # 예: 비용지원안내!$B$5:$B$42
                rm = re.match(r"(?:'?)([^'!]+)(?:'?)!\$?([A-Z]+)\$?(\d+):\$?[A-Z]+\$?(\d+)", ref)
                if rm:
                    sn, col, r1, r2 = rm.group(1), rm.group(2), int(rm.group(3)), int(rm.group(4))
                    purpose = _read_range_column(zf, sn, col, r1, r2)
            break

    return {"purpose": purpose, "payment": payment}


def get_support_limits(src_path_or_bytes):
    """
    '비용지원안내' 시트에서 목적(B)별 지원한도(F) 를 읽어 매핑을 만든다.
    반환: {목적명: 한도(int, 원) 또는 None}
          F 칸이 '승인금액' 등 정액이 아니면 None(상한 없음).
          범위(예: '5,500 ~11,000원')는 상한값(최댓값)을 사용.
    """
    raw = _read_bytes(src_path_or_bytes)
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    if SUPPORT_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SUPPORT_SHEET]

    # 지원한도(F)는 조식/중식/석식처럼 여러 목적이 한 칸을 공유(병합)한다.
    # 병합 구간의 앵커 값을 각 행으로 펼쳐 둔다.
    merged_f = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= 6 <= mr.max_col:
            anchor = ws.cell(mr.min_row, mr.min_col).value
            for rr in range(mr.min_row, mr.max_row + 1):
                merged_f[rr] = anchor

    limits = {}
    for r in range(5, ws.max_row + 1):
        purpose = ws.cell(r, 2).value          # B: 목적
        if not purpose:
            continue
        cap_text = str(merged_f.get(r) if r in merged_f else ws.cell(r, 6).value or "")
        nums = [int(x.replace(",", "")) for x in re.findall(r"[\d,]*\d", cap_text)]
        # 금액으로 볼 만한 값(>=1000)만 한도로 인정. 없으면 상한 없음(None)
        amounts = [n for n in nums if n >= 1000]
        limits[str(purpose).strip()] = max(amounts) if amounts else None
    return limits


def get_note_examples(src_path_or_bytes):
    """
    '비용지원안내' 시트에서 목적(B)별 비고작성예시(D)를 읽어 매핑을 만든다.
    반환: {목적명: 비고작성예시(str)}  (비고칸 placeholder 로 사용)
    """
    raw = _read_bytes(src_path_or_bytes)
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    if SUPPORT_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SUPPORT_SHEET]
    out = {}
    for r in range(5, ws.max_row + 1):
        purpose = ws.cell(r, 2).value          # B: 목적
        example = ws.cell(r, 4).value          # D: 비고작성예시
        if purpose and example:
            # placeholder 용 — 줄바꿈은 공백으로 정리
            out[str(purpose).strip()] = " ".join(str(example).split())
    return out


# ---------------------------------------------------------------- 검증/정렬
def _people_count(participants) -> int:
    """참여자 문자열을 콤마/공백으로 나눠 인원수를 센다. (매크로 TextToColumns 대체)"""
    s = str(participants or "").strip()
    if not s:
        return 0
    return len([p for p in re.split(r"[,\s]+", s) if p])


def _pay_key(payment) -> str:
    """결제방식 정렬 키(오름차순). '1.개인카드' 형태라 문자열 그대로 오름차순이면 1→2→3 순."""
    return str(payment or "")


def sort_for_claim(records):
    """비용청구서 정렬: 결제방식 오름차순, 영수일자 오름차순 (매크로 Sort와 동일)."""
    def key(r):
        d = _to_date(r.get("date"))
        return (_pay_key(r.get("payment")), d or date.max)
    return sorted(records, key=key)


def validate_claims(records, limits=None, year=None):
    """'동작매크로_매크로검토'의 3단계 검증을 재구현한다.

    records: list[dict]  키 = date, store, purpose, amount, payment, claim,
             participants, time, region, note
    limits : {목적명: 1인당/정액 한도(int) 또는 None}  (get_support_limits 결과)
    year   : 당해연도(int). None 이면 검증 안 함(연도 범위 체크 생략).

    반환: list[dict]  이슈 = {row(0-based), field, code, message}
    """
    limits = limits or {}
    issues = []

    # 값이 하나라도 있는(=검증 대상) 행만
    def has_any(r):
        return any(str(r.get(k) or "").strip()
                   for k in ("date", "store", "purpose", "amount", "payment", "claim"))

    rows = [(i, r) for i, r in enumerate(records) if has_any(r)]

    # 3단계용: (영수일자, 목적)별 청구금액 합계 미리 집계
    group_claim = {}
    for _, r in rows:
        d = _to_date(r.get("date"))
        amt = _to_amount(r.get("claim"))
        if d is not None and r.get("purpose") and amt:
            group_claim[(d, str(r.get("purpose")).strip())] = \
                group_claim.get((d, str(r.get("purpose")).strip()), 0) + amt

    for i, r in rows:
        purpose = str(r.get("purpose") or "").strip()

        # 1단계: 필수항목 누락
        for field, label in (("date", "영수일자"), ("store", "거래처명"),
                              ("purpose", "목적"), ("amount", "영수금액"),
                              ("payment", "결제방식"), ("claim", "청구금액")):
            if not str(r.get(field) or "").strip():
                issues.append({"row": i, "field": field, "code": "missing",
                               "message": f"{label}이(가) 비어 있습니다."})

        # 2단계: 형식/범위 (영수일자가 있을 때만)
        d = _to_date(r.get("date"))
        if str(r.get("date") or "").strip():
            if d is None:
                issues.append({"row": i, "field": "date", "code": "bad_date",
                               "message": "영수일자를 날짜로 인식할 수 없습니다."})
            elif year and not (date(year, 1, 1) <= d <= date(year, 12, 31)):
                issues.append({"row": i, "field": "date", "code": "year_out",
                               "message": f"{year}년 외의 영수일자입니다."})

        amt = _to_amount(r.get("amount"))
        if str(r.get("amount") or "").strip() and amt is None:
            issues.append({"row": i, "field": "amount", "code": "bad_amount",
                           "message": "영수금액이 숫자가 아닙니다."})
        claim = _to_amount(r.get("claim"))
        if str(r.get("claim") or "").strip() and claim is None:
            issues.append({"row": i, "field": "claim", "code": "bad_claim",
                           "message": "청구금액이 숫자가 아닙니다."})
        if amt and claim and claim > amt:
            issues.append({"row": i, "field": "claim", "code": "claim_gt_amount",
                           "message": "청구금액이 영수금액보다 큽니다."})

        # 3단계: 식대 1인당 한도 (참여자 인원수 × 한도)
        if purpose in MEAL_PURPOSES:
            count = _people_count(r.get("participants"))
            if count == 0:
                issues.append({"row": i, "field": "participants", "code": "meal_no_participant",
                               "message": "식대는 참여자를 반드시 기재해야 합니다."})
            else:
                cap = limits.get(purpose)
                if cap is not None and d is not None:
                    total = group_claim.get((d, purpose), 0)
                    if total > cap * count:
                        issues.append({"row": i, "field": "claim", "code": "over_limit",
                                       "message": (f"동일 일자·목적 청구금액 합계({total:,}원)가 "
                                                   f"1인당 한도×인원({cap:,}×{count})을 초과했습니다.")})

    return issues


def _fill_claim_sheet_xml(claim_xml, records_sorted, date_as_serial=False):
    """비용청구서 시트 XML에 정렬된 데이터를 17행부터 매핑 기입한다.
    작성시트→비용청구서 열 매핑(매크로와 동일):
      C→C 영수일자, D→D 거래처, E→E 목적, participants→F 참여자, note→G 비고,
      payment→J 결제방식, amount→K 영수금액, claim→L 청구금액, time→N 시간, region→O 지역
    지급금액(P)·사유(Q)는 경영지원팀 기재란이라 비운다. H4=작성일, I3/I9=완료 상태.
    date_as_serial=True 면 영수일자를 엑셀 날짜 일련번호(숫자)로 기입해 셀의 날짜서식이
    그대로 적용되게 한다(최종 .xlsx용). False 면 텍스트(매크로 문자열 비교용).
    """
    n = len(records_sorted)
    if n > CLAIM_MAX_ROWS:
        raise ValueError(f"청구내역이 {n}건으로 양식 최대({CLAIM_MAX_ROWS}건)를 초과합니다. "
                         f"현재는 {CLAIM_MAX_ROWS}건까지 지원합니다.")
    for i, rec in enumerate(records_sorted):
        r = CLAIM_FIRST_ROW + i
        d = _to_date(rec.get("date"))
        if d is not None:
            if date_as_serial:
                claim_xml = _set_cell(claim_xml, f"C{r}", "num", _date_serial(d))
            else:
                claim_xml = _set_cell(claim_xml, f"C{r}", "str", d.isoformat())
        store = rec.get("store")
        if store:
            claim_xml = _set_cell(claim_xml, f"D{r}", "str", str(store).strip())
        purpose = rec.get("purpose")
        if purpose:
            claim_xml = _set_cell(claim_xml, f"E{r}", "str", str(purpose).strip())
        participants = rec.get("participants")
        if participants:
            claim_xml = _set_cell(claim_xml, f"F{r}", "str", str(participants).strip())
        note = rec.get("note")
        if note:
            claim_xml = _set_cell(claim_xml, f"G{r}", "str", str(note).strip())
        payment = rec.get("payment")
        if payment:
            claim_xml = _set_cell(claim_xml, f"J{r}", "str", str(payment).strip())
        amt = _to_amount(rec.get("amount"))
        if amt is not None:
            claim_xml = _set_cell(claim_xml, f"K{r}", "num", amt)
        claim = _to_amount(rec.get("claim"))
        if claim is not None:
            claim_xml = _set_cell(claim_xml, f"L{r}", "num", claim)
        t = _to_time(rec.get("time"))
        if t is not None:
            claim_xml = _set_cell(claim_xml, f"N{r}", "num", repr(_time_fraction(t)))
        region = rec.get("region")
        if region:
            claim_xml = _set_cell(claim_xml, f"O{r}", "str", str(region).strip())
    # 작성일(H4)·검토/서명 상태(I3/I9) — 매크로 '비용청구서생성'과 동일하게 완료로 표기
    claim_xml = _set_cell(claim_xml, "H4", "str", date.today().isoformat())
    claim_xml = _set_cell(claim_xml, "I3", "str", "서명완료")
    claim_xml = _set_cell(claim_xml, "I9", "str", "완료")
    return claim_xml


# 최종 .xlsx 에 남길 시트(매크로 '비용청구서생성' 결과와 동일: 비용청구서 + 교통비상세)
CLAIM_KEEP_SHEETS = ["비용청구서", "교통비상세"]


def _claim_summary_cache(records):
    """비용청구서 상단 합계 수식(H10~H13)의 캐시값을 미리 계산한다.
    H10=SUM(K:K)영수금액합계, H11=SUMIF 개인형법인카드분, H12=SUM(L:L)청구금액합계, H13=지급금액(0).
    (수식 <f>는 유지하되 <v> 캐시를 채워, Excel이 아닌 뷰어에서도 합계가 보이게 한다.)"""
    amt = sum(_to_amount(r.get("amount")) or 0 for r in records)
    claim = sum(_to_amount(r.get("claim")) or 0 for r in records)
    card = sum((_to_amount(r.get("amount")) or 0) for r in records
               if "개인형법인카드" in str(r.get("payment") or ""))
    return {"H10": amt, "H11": card, "H12": claim, "H13": 0}


def _patch_formula_cache(xml, coord, value):
    """수식 셀(coord)의 <v> 캐시값만 교체한다(<f> 수식은 그대로)."""
    m = re.search(r'<c r="%s"((?:\s+[^>]*?)?)>(.*?)</c>' % re.escape(coord), xml, re.S)
    if not m or "<f" not in m.group(2):
        return xml
    inner = m.group(2)
    if "<v>" in inner:
        inner = re.sub(r"<v>.*?</v>", f"<v>{value}</v>", inner, count=1, flags=re.S)
    else:
        inner = inner + f"<v>{value}</v>"
    return xml[:m.start(2)] + inner + xml[m.end(2):]


def build_claim_xlsx(src_path_or_bytes, records_sorted, basic_info=None):
    """매크로 '동작매크로_비용청구서생성'처럼, 비용청구서(+교통비상세) 시트만 담은 독립 .xlsx 를 만든다.

    zip 수술 방식 — 나머지 시트/매크로/도형/calcChain 만 제거하고, 남기는 시트의 XML·스타일·
    조건부서식·병합·표시형식·인쇄영역은 원본 그대로 보존한다. 비용청구서 시트에만 데이터·상단
    정보·합계 캐시를 기입한다. records_sorted 는 sort_for_claim 으로 정렬해 넘길 것.
    반환: (BytesIO, count).  25건 초과면 ValueError.
    """
    import posixpath
    n = len(records_sorted)
    if n > CLAIM_MAX_ROWS:
        raise ValueError(f"청구내역이 {n}건으로 양식 최대({CLAIM_MAX_ROWS}건)를 초과합니다. "
                         f"현재는 {CLAIM_MAX_ROWS}건까지 지원합니다.")
    raw = _read_bytes(src_path_or_bytes)
    zin = zipfile.ZipFile(BytesIO(raw))
    names = set(zin.namelist())

    wbxml = zin.read("xl/workbook.xml").decode("utf-8")
    relsxml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rid_target = {m.group(1): m.group(2) for m in
                  re.finditer(r'<Relationship [^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', relsxml)}

    all_sheets = []
    for m in re.finditer(r"<sheet\b([^>]*?)/>", wbxml):
        a = m.group(1)
        all_sheets.append({
            "name": re.search(r'name="([^"]+)"', a).group(1),
            "sheetId": re.search(r'sheetId="([^"]+)"', a).group(1),
            "rid": re.search(r'r:id="([^"]+)"', a).group(1),
        })
    orig_order = [s["name"] for s in all_sheets]
    keep = [s for name in CLAIM_KEEP_SHEETS for s in all_sheets if s["name"] == name]
    if not any(s["name"] == CLAIM_SHEET for s in keep):
        raise ValueError(f"'{CLAIM_SHEET}' 시트를 찾을 수 없습니다.")
    new_order = [s["name"] for s in keep]

    def wpath(rid):
        t = rid_target[rid].lstrip("/")
        return t if t.startswith("xl/") else "xl/" + t
    keep_ws_path = {s["name"]: wpath(s["rid"]) for s in keep}
    kept_ws_targets = set(keep_ws_path.values())

    # 남길 워크시트의 rels 가 참조하는 추가 파트(printerSettings 등) 수집
    ws_rels_files, extra_parts = {}, set()
    for wp in keep_ws_path.values():
        d, base = wp.rsplit("/", 1)
        rels_name = f"{d}/_rels/{base}.rels"
        if rels_name in names:
            rtext = zin.read(rels_name).decode("utf-8")
            ws_rels_files[rels_name] = rtext
            for mm in re.finditer(r'Target="([^"]+)"', rtext):
                extra_parts.add(posixpath.normpath(posixpath.join(d, mm.group(1))))

    # ---- 비용청구서 시트 XML 편집: 상단정보 + 데이터 + 합계 캐시 ----
    claim_path = keep_ws_path[CLAIM_SHEET]
    claim_xml = zin.read(claim_path).decode("utf-8")
    bi = basic_info or {}
    claim_xml = _set_cell(claim_xml, "H5", "str", str(bi.get("dept", "") or "").strip())
    claim_xml = _set_cell(claim_xml, "H6", "str", str(bi.get("name", "") or "").strip())
    claim_xml = _set_cell(claim_xml, "H7", "str", str(bi.get("card", "") or "").strip())
    title = str(bi.get("title", "") or "").strip()
    if title:
        claim_xml = _set_cell(claim_xml, "B2", "str", f"{title}청구서")
    claim_xml = _fill_claim_sheet_xml(claim_xml, records_sorted, date_as_serial=True)  # 데이터 + H4/I3/I9
    for coord, val in _claim_summary_cache(records_sorted).items():
        claim_xml = _patch_formula_cache(claim_xml, coord, val)
    # 파일을 열면 비용청구서 시트가 먼저 보이도록 선택 상태로 만든다.
    if "tabSelected" not in re.search(r"<sheetView\b[^>]*>", claim_xml).group(0):
        claim_xml = re.sub(r"(<sheetView\b)([^>]*?)(>)",
                           r'\1\2 tabSelected="1"\3', claim_xml, count=1)

    # ---- workbook.xml: 시트 목록·정의된이름·재계산 플래그 정리 ----
    sheets_xml = "<sheets>" + "".join(
        f'<sheet name="{s["name"]}" sheetId="{s["sheetId"]}" r:id="{s["rid"]}"/>'
        for s in keep) + "</sheets>"
    wbxml = re.sub(r"<sheets>.*?</sheets>", lambda _m: sheets_xml, wbxml, count=1, flags=re.S)
    dn = re.search(r"<definedNames>.*?</definedNames>", wbxml, re.S)
    if dn:
        out_names = []
        for dm in re.finditer(r"<definedName\b[^>]*>.*?</definedName>", dn.group(0), re.S):
            whole = dm.group(0)
            lsi = re.search(r'localSheetId="(\d+)"', whole)
            if lsi:
                oname = orig_order[int(lsi.group(1))] if int(lsi.group(1)) < len(orig_order) else None
                if oname not in new_order:
                    continue  # 제거된 시트용 정의된 이름은 버림
                whole = re.sub(r'localSheetId="\d+"', f'localSheetId="{new_order.index(oname)}"', whole)
            out_names.append(whole)
        new_dn = ("<definedNames>" + "".join(out_names) + "</definedNames>") if out_names else ""
        wbxml = wbxml[:dn.start()] + new_dn + wbxml[dn.end():]
    if "fullCalcOnLoad" not in wbxml:
        wbxml = re.sub(r"<calcPr\b([^>]*?)/>",
                       lambda m: f"<calcPr{m.group(1)} fullCalcOnLoad=\"1\"/>", wbxml, count=1)
    # 열었을 때 첫 시트(비용청구서)가 활성화되도록 activeTab 을 0 으로.
    if re.search(r"<workbookView\b[^>]*>", wbxml):
        if "activeTab=" in re.search(r"<workbookView\b[^>]*>", wbxml).group(0):
            wbxml = re.sub(r'(<workbookView\b[^>]*?)\s*activeTab="\d+"',
                           r"\1", wbxml, count=1)
        wbxml = re.sub(r"(<workbookView\b)([^>]*?)(/?>)",
                       r'\1\2 activeTab="0"\3', wbxml, count=1)

    # ---- workbook.xml.rels: 남긴 워크시트/스타일/테마/문자열만 유지 ----
    def keep_rel(m):
        whole = m.group(0)
        typ = re.search(r'Type="([^"]+)"', whole).group(1)
        tgt = re.search(r'Target="([^"]+)"', whole).group(1)
        tnorm = posixpath.normpath(posixpath.join("xl", tgt.lstrip("/")))
        if typ.endswith("/worksheet"):
            return whole if tnorm in kept_ws_targets else ""
        if typ.endswith("/calcChain") or typ.endswith("/vbaProject"):
            return ""
        return whole
    new_rels = re.sub(r"<Relationship\b[^>]*/>", keep_rel, relsxml)

    # ---- 출력 파트 목록 확정 ----
    base_parts = ["_rels/.rels", "docProps/core.xml", "docProps/app.xml",
                  "xl/styles.xml", "xl/sharedStrings.xml", "xl/theme/theme1.xml"]
    included = {"xl/workbook.xml", "xl/_rels/workbook.xml.rels"}
    included.update(p for p in base_parts if p in names)
    included.update(keep_ws_path.values())
    included.update(ws_rels_files.keys())
    included.update(p for p in extra_parts if p in names)

    # ---- [Content_Types].xml: 포함 파트의 Override만 유지 + 워크북 타입 xlsx 로 ----
    ct = zin.read("[Content_Types].xml").decode("utf-8")
    XLSX_MAIN = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"

    def keep_override(m):
        whole = m.group(0)
        part = re.search(r'PartName="/([^"]+)"', whole).group(1)
        if part.lower() not in {p.lower() for p in included}:
            return ""
        if part == "xl/workbook.xml":
            return re.sub(r'ContentType="[^"]+"', f'ContentType="{XLSX_MAIN}"', whole)
        return whole
    ct = re.sub(r"<Override\b[^>]*/>", keep_override, ct)

    # ---- 새 .xlsx 패키지 작성 ----
    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zo:
        zo.writestr("[Content_Types].xml", ct)
        zo.writestr("xl/workbook.xml", wbxml)
        zo.writestr("xl/_rels/workbook.xml.rels", new_rels)
        for p in base_parts:
            if p in names:
                zo.writestr(p, zin.read(p))
        for s in keep:
            wp = keep_ws_path[s["name"]]
            zo.writestr(wp, claim_xml if s["name"] == CLAIM_SHEET else
                        zin.read(wp).decode("utf-8"))
        for rn, rt in ws_rels_files.items():
            zo.writestr(rn, rt)
        for ep in extra_parts:
            if ep in names:
                zo.writestr(ep, zin.read(ep))
    zin.close()
    out.seek(0)
    return out, n


# ---------------------------------------------------------------- 메인
def fill_workbook(src_path_or_bytes, records, append=True, basic_info=None,
                  claim_data=None):
    """
    영수증 레코드를 워크북에 채워 BytesIO 로 반환한다.
    원본의 도형/이미지/매크로/서식을 100% 보존한다.

    records: list[dict]  키 = date, store, purpose, amount, payment, time,
             claim_amount(청구금액 H), region(지역 K), participants(참여자 L), note(비고 M)
    append : True 면 기존 데이터 다음 빈 행부터, False 면 15행부터
    basic_info: dict  '기초정보입력' 매크로가 비용청구서 시트에 채우는 값을 대신 기록.
             dept(소속 H5), name(성명 H6), card(법인카드번호 H7),
             title(청구 항목명 -> B2 "{title}청구서"). name 이 있으면 본인확인(I3)을
             '서명완료', 매크로검토(I9)를 '필요' 로 세팅해 엑셀에서 매크로검토만 누르면 되게 한다.
    claim_data: list[dict] | None  주어지면 매크로 '비용청구생성공통단계'를 대신해
             비용청구서 시트(17행~)에 정렬·매핑된 데이터를 직접 기입한다.
             (호출 전 sort_for_claim 으로 정렬해서 넘길 것.) 25건 초과면 ValueError.
    반환    : (BytesIO, start_row, count)
    """
    raw = _read_bytes(src_path_or_bytes)
    zin = zipfile.ZipFile(BytesIO(raw))
    sheet_path = _sheet_path_for(zin, SHEET_NAME)
    xml = zin.read(sheet_path).decode("utf-8")

    # 기초정보 + 서명/상태 + (선택)최종 청구서 데이터 (비용청구서 시트)
    bi = basic_info or {}
    claim_path, claim_xml = None, None
    need_claim = any(bi.get(k) for k in ("dept", "name", "card", "title")) \
        or claim_data is not None
    if need_claim:
        claim_path = _sheet_path_for(zin, CLAIM_SHEET)
        claim_xml = zin.read(claim_path).decode("utf-8")
    if any(bi.get(k) for k in ("dept", "name", "card", "title")):
        # 매크로 '기초정보입력' 대체
        claim_xml = _set_cell(claim_xml, "H5", "str", str(bi.get("dept", "") or "").strip())
        claim_xml = _set_cell(claim_xml, "H6", "str", str(bi.get("name", "") or "").strip())
        claim_xml = _set_cell(claim_xml, "H7", "str", str(bi.get("card", "") or "").strip())
        title = str(bi.get("title", "") or "").strip()
        if title:
            claim_xml = _set_cell(claim_xml, "B2", "str", f"{title}청구서")
        # 본인확인(I3)·매크로검토(I9) 상태: 이름이 있으면 서명완료 처리
        if str(bi.get("name", "") or "").strip():
            claim_xml = _set_cell(claim_xml, "I3", "str", "서명완료")
        claim_xml = _set_cell(claim_xml, "I9", "str", "필요")
    if claim_data is not None:
        # 매크로 '비용청구생성공통단계' 대체 — 정렬된 데이터를 비용청구서 시트에 매핑 기입
        claim_xml = _fill_claim_sheet_xml(claim_xml, claim_data)

    start = _first_empty_row(xml) if append else FIRST_DATA_ROW

    for i, rec in enumerate(records):
        r = start + i
        d = _to_date(rec.get("date"))
        if d is not None:
            # 매크로가 영수일자를 문자열로 비교(< "2026-01-01")하므로 YYYY-MM-DD 텍스트로 기록
            xml = _set_cell(xml, f"C{r}", "str", d.isoformat())
        store = rec.get("store")
        if store:
            xml = _set_cell(xml, f"D{r}", "str", str(store).strip())
        purpose = rec.get("purpose")
        if purpose:
            xml = _set_cell(xml, f"E{r}", "str", str(purpose).strip())
        payment = rec.get("payment")
        if payment:
            xml = _set_cell(xml, f"G{r}", "str", str(payment).strip())
        amt = _to_amount(rec.get("amount"))
        if amt is not None:
            xml = _set_cell(xml, f"F{r}", "num", amt)
        t = _to_time(rec.get("time"))
        if t is not None:
            xml = _set_cell(xml, f"J{r}", "num", repr(_time_fraction(t)))
        # 추가 입력칸: 청구금액(H)·지역(K)·참여자(L)·비고(M)
        claim = _to_amount(rec.get("claim_amount"))
        if claim is not None:
            xml = _set_cell(xml, f"H{r}", "num", claim)
        region = rec.get("region")
        if region:
            xml = _set_cell(xml, f"K{r}", "str", str(region).strip())
        participants = rec.get("participants")
        if participants:
            xml = _set_cell(xml, f"L{r}", "str", str(participants).strip())
        note = rec.get("note")
        if note:
            xml = _set_cell(xml, f"M{r}", "str", str(note).strip())

    # 새 zip 작성 (대상 시트만 교체, 나머지 원본 그대로 복사)
    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                data = xml.encode("utf-8")
            elif claim_path and item.filename == claim_path:
                data = claim_xml.encode("utf-8")
            elif item.filename == "xl/workbook.xml":
                # 목적(E) 기반 안내 수식(I·N) 등이 열 때 재계산되도록
                data = _force_full_recalc(data.decode("utf-8")).encode("utf-8")
            # 압축 방식/속성 보존
            zi = zipfile.ZipInfo(item.filename, date_time=item.date_time)
            zi.compress_type = item.compress_type
            zi.external_attr = item.external_attr
            zi.internal_attr = item.internal_attr
            zi.create_system = item.create_system
            zout.writestr(zi, data)
    zin.close()
    out.seek(0)
    return out, start, len(records)


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "비용청구_남소희_20260604.xlsm"
    opts = get_dropdown_options(path)
    print("목적 옵션:", len(opts["purpose"]), opts["purpose"][:3], "...")
    print("결제방식 옵션:", opts["payment"])
    demo = [
        {"date": "2026-06-01", "store": "테스트상회", "purpose": "야근식비",
         "amount": "12,500", "payment": "1.개인카드", "time": "19:30"},
        {"date": "2026.06.02", "store": "분식나라 <김밥>", "purpose": "기타식비",
         "amount": 8000, "payment": "2.현금", "time": "20:05"},
    ]
    buf, start, n = fill_workbook(path, demo, append=True)
    with open("test_output.xlsm", "wb") as f:
        f.write(buf.read())
    print(f"채움 완료: {start}행부터 {n}건 -> test_output.xlsm")
