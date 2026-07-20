"""
overtime_filler.py
월간 근태현황(.xlsx)을 읽어 '초과근무(수당)신청서' 양식을 채운다.

규칙 (사용자 지정)
  - 포함 대상: '승인 초과 근로시간' > 0 인 날만.
  - 근무시작(I) = 출근시간 + 9시간   (표준근무 8h + 점심 1h)
                  예) 08:00 출근 -> 17:00, 09:12 출근 -> 18:12
                  단, 08:00 이전 조기출근은 08:00부터 근무한 것으로 보아
                  근무시작을 17:00로 노출한다. 예) 06:43 출근 -> 17:00
                  (출근 원본값은 그대로 두고, 근무시작 기준만 08:00로 하한)
  - 근무종료(J) = 퇴근시간
  - 근무시간(K) = 양식 수식이 J-I 로 자동 계산 (0.5시간 단위)
  - 실 근무시작(L) = 근무시작(I), 실 근무종료(M) = 근무종료(J)

★ 도형 보존 ★
양식에는 결재칸 등 도형/VML 이 있어, openpyxl 재저장 시 사라진다.
그래서 .xlsx 를 zip 으로 열어 시트 XML 의 대상 셀 값만 직접 교체한다.
근무시간(K)·신청시간(S) 등 수식과 날짜 자동생성(C2 기반)은 그대로 둔다.
"""

import re
import math
import zipfile
from datetime import date
from io import BytesIO

import openpyxl

# excel_filler 의 범용 zip/XML 헬퍼 재사용
from excel_filler import (
    _read_bytes, _sheet_path_for, _set_cell, _force_full_recalc, _date_serial,
)


def _set_formula_cache(xml, coord, value):
    """수식 셀(<f>)은 그대로 두고 캐시값(<v>)만 설정/교체한다.
    다운로드 파일을 '제한된 보기'로 열면 Excel이 수식을 재계산하지 않아 신청시간이
    0(저장된 캐시값)으로 보인다 → 미리 계산한 값을 캐시로 넣어 바로 보이게 한다.
    ('편집 사용'을 누르면 어차피 재계산되어 자동 보정되므로 안전.)
    수식이 없는 셀이면 그대로 둔다."""
    m = re.search(r'<c r="%s"((?:\s+[^>]*?)?)(/>|>.*?</c>)' % re.escape(coord), xml, re.S)
    if not m:
        return xml
    attrs, body = m.group(1), m.group(2)
    fm = re.search(r'<f\b[^>]*>.*?</f>|<f\b[^>]*/>', body, re.S)
    if not fm:
        return xml  # 수식 셀이 아니면 건드리지 않음
    new_attrs = re.sub(r'\s+t="[^"]*"', '', attrs)   # 결과는 숫자 → t 속성 제거
    new_cell = f'<c r="{coord}"{new_attrs}>{fm.group(0)}<v>{value}</v></c>'
    return xml[:m.start()] + new_cell + xml[m.end():]


def _nf(x):
    """캐시값 숫자 포맷: 0 / 2.5 / 5 처럼 깔끔하게."""
    return "%g" % round(float(x), 4)


def _to_hours(s):
    """'1' / '1.5' / '1:30' → 소수 시간(float). 빈값/이상값 → 0.0.
    양식 N(제외할 시간)·Q(대체휴무 시간)는 '시간 숫자'를 빼므로 항상 시간 단위로 저장한다."""
    s = str(s or "").strip()
    if not s:
        return 0.0
    if ":" in s:
        sec = _parse_hms(s)
        return (sec / 3600.0) if sec is not None else 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

FORM_SHEET = "양식"
STANDARD_WORK_SECONDS = 9 * 3600   # 정규근무 9시간(점심 포함)
# 미승인 초과를 신청시간에 합치는 상한: 반올림으로 떨어진 '초 단위 잔여분'만 인정한다.
# 1분(60초) 이상이면 실제 미승인 근무로 보고 신청시간에서 제외한다(과다 청구 방지).
UNAPPROVED_CARRY_MAX_SECONDS = 60
WORK_START_FLOOR_SECONDS = 8 * 3600  # 근무시작 기준 하한: 08:00 이전 출근은 08:00부터로 본다
DAY_SECONDS = 86400


# ---------------------------------------------------------------- 시간 파싱
def _parse_hms(v):
    """'HH:MM:SS' 또는 time -> 총 초(int). 빈값/0 이면 None 또는 0."""
    if v is None or v == "":
        return None
    if hasattr(v, "hour"):  # datetime.time / datetime
        return v.hour * 3600 + v.minute * 60 + getattr(v, "second", 0)
    s = str(v).strip()
    parts = s.split(":")
    try:
        parts = [int(p) for p in parts]
    except ValueError:
        return None
    while len(parts) < 3:
        parts.append(0)
    h, m, sec = parts[:3]
    return h * 3600 + m * 60 + sec


def _fraction(seconds):
    return seconds / DAY_SECONDS


def _set_hours_cell(xml, ref, val):
    """대체휴무시간 칸: 'HH:MM' 이면 시간값(분수), 숫자면 숫자, 그 외엔 문자열로 기록.
    빈값이면 그대로 둔다."""
    if val is None:
        return xml
    s = str(val).strip()
    if s == "":
        return xml
    if ":" in s:
        sec = _parse_hms(s)
        if sec is not None:
            return _set_cell(xml, ref, "num", repr(_fraction(sec)))
    try:
        return _set_cell(xml, ref, "num", repr(float(s)))
    except ValueError:
        return _set_cell(xml, ref, "str", s)


def _find_header_row(ws, max_scan=15):
    """'일자' 헤더가 있는 행을 찾는다. 양식별로 헤더 행 위치가 다를 수 있다."""
    for r in range(1, max_scan + 1):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(r, c).value or "").strip() == "일자":
                return r
    return 7  # 못 찾으면 구양식 기준으로 폴백


def _find_ot_column(ws, header_row):
    """'승인 초과 근로시간' 합계 열을 찾는다.
    - 구양식: 헤더행에 '승인 초과 근로시간' 단일 컬럼.
    - 신양식: 상위 그룹행(헤더행-1)에 '승인' 병합헤더 + 헤더행에 '초과근로시간' 합계 컬럼
              (그 옆 연장/야간/휴일 등은 세부 내역이라 다른 값이므로 제외).
    못 찾으면 None (호출부에서 명확히 에러 처리)."""
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(header_row, c).value or "").strip() == "승인 초과 근로시간":
            return c
    if header_row > 1:
        for c in range(1, ws.max_column + 1):
            grp = str(ws.cell(header_row - 1, c).value or "").strip()
            sub = str(ws.cell(header_row, c).value or "").strip()
            if grp == "승인" and sub in ("초과근로시간", "초과 근로시간"):
                return c
    return None


def _find_unapproved_ot_column(ws, header_row):
    """'미승인 초과 근로시간' 합계 열을 찾는다.
    - 헤더행에 '미승인 초과 근로시간' 단일 컬럼이 있거나,
    - 상위 그룹행(헤더행-1)에 '미승인' 병합헤더 + 헤더행에 '초과근로시간' 합계 컬럼.
    못 찾으면 사용자 지정 위치인 U열(21)로 폴백. 그마저 없으면 None."""
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(header_row, c).value or "").strip() == "미승인 초과 근로시간":
            return c
    if header_row > 1:
        for c in range(1, ws.max_column + 1):
            grp = str(ws.cell(header_row - 1, c).value or "").strip()
            sub = str(ws.cell(header_row, c).value or "").strip()
            if grp == "미승인" and sub in ("초과근로시간", "초과 근로시간"):
                return c
    return 21 if ws.max_column >= 21 else None  # U열(21) 폴백


# ---------------------------------------------------------------- 근태 읽기
def parse_attendance(src_path_or_bytes):
    """
    월간 근태현황을 읽어 (name, year, month, records, unapproved_total_h) 반환.
    records: list[dict] 키 = day(int), clock_in(sec), clock_out(sec),
             approved_ot(sec)  — 승인초과>0 인 날만.
    unapproved_total_h: 월 전체 '미승인 초과 근로시간' 합계(시간). 화면 지표용.
             승인 신청시간과 동일하게 일자별 0.5시간 단위 내림 후 합산한다.
    """
    raw = _read_bytes(src_path_or_bytes)
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    ws = wb.active

    # 이름: 'B2 = "남소희 월간 근태현황"'
    title = str(ws["B2"].value or "").strip()
    name = title.split()[0] if title else ""

    # 조회기간: 'C5 = "202605"'
    period = str(ws["C5"].value or "").strip()
    year, month = None, None
    if len(period) >= 6 and period[:6].isdigit():
        year, month = int(period[:4]), int(period[4:6])

    # 헤더 행은 양식마다 다를 수 있어 '일자' 텍스트로 찾는다.
    # 열: B(일자) C(출근) F(퇴근) + 승인 초과 근로시간(구양식 단일열/신양식 그룹+합계열)
    header_row = _find_header_row(ws)
    col = {c.value: c.column for c in ws[header_row] if c.value}
    c_date = col.get("일자", 2)
    c_in = col.get("출근시간", 3)
    c_out = col.get("퇴근시간", 6)
    c_ot = _find_ot_column(ws, header_row)
    if c_ot is None:
        raise ValueError("근태현황 양식에서 '승인 초과 근로시간' 열을 찾을 수 없습니다.")
    c_unot = _find_unapproved_ot_column(ws, header_row)  # 미승인 초과 근로시간(U열)

    records = []
    # 월 전체 미승인 초과 근로시간 합계(시간) — 실제 일자 행만, 승인 신청시간과 동일하게
    # 일자별 0.5시간 단위 내림 후 합산(예: 2:14→2.0, 0:41→0.0).
    unapproved_total = 0.0
    for r in range(header_row + 1, ws.max_row + 1):
        dval = ws.cell(r, c_date).value
        if not dval:
            continue
        # 일자(day) 추출 — 날짜가 아닌 행(맨 아래 '근로시간 합계' 총계 행 등)은 건너뛴다.
        # (총계 행에도 미승인 합계값이 들어 있어, 안 거르면 일자별 합과 중복돼 2배가 된다.)
        ds = str(dval)
        day = None
        if hasattr(dval, "day"):
            day = dval.day
        else:
            m = re.search(r"-(\d{2})$", ds) or re.search(r"-(\d{1,2})\b", ds)
            if m:
                day = int(m.group(1))
        if day is None:
            continue
        # 미승인 초과 근로시간(초). 승인 유무와 무관하게 (실제 일자 행이면) 지표에 누적하고,
        # 각 날짜 레코드에도 저장한다(신청시간 계산 시 승인초과에 합쳐 초 단위 잔여분을 살린다).
        # 지표 합계는 승인 신청시간 규칙과 동일하게 일자별 0.5시간 단위 내림 후 합산.
        un_sec = 0
        if c_unot:
            un_sec = _parse_hms(ws.cell(r, c_unot).value) or 0
            unapproved_total += math.floor(un_sec / 3600.0 * 2) / 2
        ot = _parse_hms(ws.cell(r, c_ot).value) or 0
        cin = _parse_hms(ws.cell(r, c_in).value)
        cout = _parse_hms(ws.cell(r, c_out).value)
        if ot <= 0 or cin is None or cout is None:
            continue
        if year is None and hasattr(dval, "year"):
            year, month = dval.year, dval.month
        records.append({"day": day, "clock_in": cin, "clock_out": cout,
                        "approved_ot": ot, "unapproved_ot": un_sec})
    return name, year, month, records, unapproved_total


# ---------------------------------------------------------------- 양식 채우기
def fill_overtime(template_path_or_bytes, attendance_path_or_bytes,
                  name=None, month=None, extras=None, dept_position=None):
    """
    근태현황을 읽어 초과근무신청서 양식을 채워 (BytesIO, count) 반환.
    name/month 를 직접 주면 근태 파일 값보다 우선한다.
    dept_position 을 주면 '부서명 / 직위' 칸(D7)에 채운다.

    extras: dict[int day] -> {"payoff": "O"/"X", "hours": "HH:MM"|숫자, "note": str}
            사용자가 표에서 고른 대체휴무지급(P)·대체휴무시간(Q)·비고(R) 값.
            (헤더는 15행, 일자별 데이터는 16+일자 행)
    """
    extras = extras or {}
    a_name, a_year, a_month, records, _ = parse_attendance(attendance_path_or_bytes)
    name = name or a_name
    month = month or a_month

    raw = _read_bytes(template_path_or_bytes)
    zin = zipfile.ZipFile(BytesIO(raw))
    sheet_path = _sheet_path_for(zin, FORM_SHEET)
    xml = zin.read(sheet_path).decode("utf-8")

    # 기본정보: 월(C2), 작성일(D6=TODAY()), 부서명/직위(D7), 성명(D8)
    if month:
        xml = _set_cell(xml, "C2", "num", int(month))
    # D6는 =TODAY() 수식 셀. '제한된 보기'(편집 사용 전)는 재계산을 안 하고 양식에
    # 저장돼 있던 캐시값(예: 템플릿을 마지막으로 연 날짜)을 그대로 보여준다.
    # 캐시를 생성 시점(오늘)으로 갱신해, 편집 사용 전에도 실제 작성일이 보이게 한다.
    xml = _set_formula_cache(xml, "D6", str(_date_serial(date.today())))
    # 수행일자(C열)도 =DATE(YEAR($D$6),$C$2,$B{행}) 수식이라, 편집 사용 전에는
    # 캐시된 옛 달(예: 2월) 날짜가 그대로 보인다. 새 달(C2)·올해(D6) 기준으로 캐시를 갱신한다.
    # serial = (해당 월 1일 serial) + (일-1) → 엑셀 DATE의 월 넘김 규칙과 동일하게 계산됨.
    if month:
        base = _date_serial(date(date.today().year, int(month), 1))
        for day in range(1, 32):        # 1일=17행, day -> 16+day (없거나 수식 아니면 자동 무시)
            xml = _set_formula_cache(xml, f"C{16 + day}", str(base + (day - 1)))
    if dept_position:
        xml = _set_cell(xml, "D7", "str", str(dept_position).strip())
    if name:
        xml = _set_cell(xml, "D8", "str", name)

    # 일자별 행 채우기 (양식: 1일=17행, day -> 16+day)
    total_s = 0.0   # 신청시간 합계(S12) 캐시값
    for rec in records:
        day = rec["day"]
        r = 16 + day
        # 근무시작(I) = 출근+9h. 단 08:00 이전 조기출근은 08:00부터 근무한 것으로 보아
        # 근무시작을 하한 처리한다(예: 06:43 출근 -> 08:00 기준 -> 17:00). 출근 원본값은 유지.
        eff_in = max(rec["clock_in"], WORK_START_FLOOR_SECONDS)
        i_sec = eff_in + STANDARD_WORK_SECONDS            # 근무시작 = max(출근,08:00)+9h
        j_sec = rec["clock_out"]                          # 근무종료 = 퇴근(실제)
        ot_sec = rec.get("approved_ot", 0) or 0           # 승인 초과 근로시간
        un_raw = rec.get("unapproved_ot", 0) or 0         # 미승인 초과 근로시간(초)
        # 안전장치: 1분 이상 미승인은 실제 미승인 근무로 보고 제외, 60초 미만(반올림 잔여분)만 합친다.
        un_sec = un_raw if un_raw < UNAPPROVED_CARRY_MAX_SECONDS else 0
        early_in = rec["clock_in"] < WORK_START_FLOOR_SECONDS  # 08:00 이전 조기출근
        if early_in:
            # 조기출근: 근무시작(17:00) 이전 시간은 연장에서 제외한다.
            # 신청 기준 = 퇴근 - 근무시작(=승인+미승인 전체). 실근무종료(M)=퇴근으로 두면
            # 양식 수식 S = (M-L)*24 - N - Q = (퇴근-근무시작) - 제외 - 대체휴무 가 된다.
            claim_sec = max(0, j_sec - i_sec)
            m_sec = j_sec
        else:
            # 08:00 이후 출근: 승인초과 + 미승인초과 를 신청 기준으로 한다.
            # (승인초과가 59:19처럼 정시 직전이고 남은 초가 미승인으로 떨어진 경우,
            #  실제 1시간을 넘긴 근무가 0.5로 깎이지 않도록 초 단위 잔여분을 합친다.)
            # 실근무종료(M)=근무시작+승인+미승인 → 양식 수식 S = (승인+미승인) - 제외 - 대체휴무.
            claim_sec = ot_sec + un_sec
            m_sec = i_sec + ot_sec + un_sec
        xml = _set_cell(xml, f"I{r}", "num", repr(_fraction(i_sec)))   # 근무시작(표시)
        xml = _set_cell(xml, f"J{r}", "num", repr(_fraction(j_sec)))   # 근무종료(표시=퇴근)
        xml = _set_cell(xml, f"L{r}", "num", repr(_fraction(i_sec)))   # 실 근무시작
        xml = _set_cell(xml, f"M{r}", "num", repr(_fraction(m_sec)))   # 실 근무종료 = 근무시작+승인초과

        # 사용자가 표에서 고른 값: 대체휴무지급(P)·대체휴무시간(Q)·비고(R)
        # 신청시간(S) 수식 = ROUNDDOWN(MAX(0,(근무시간) - N - Q)*2)/2 이고,
        # P 가 안내문('입력하세요')이면 숫자 대신 안내문이 나온다.
        #   X(대체휴무 미지급) -> P="X", Q 비움  -> 근무시간 전체가 신청시간에 기록
        #   O(대체휴무 지급)   -> P="O", Q=대체휴무시간 -> 그만큼 신청시간에서 차감
        ex = extras.get(day) or {}
        payoff = str(ex.get("payoff", "") or "").strip().upper()
        if payoff not in ("O", "X"):
            payoff = "X"  # 기본: 대체휴무 미지급 → 전체 신청
        xml = _set_cell(xml, f"P{r}", "str", payoff)

        # 대체휴무 시간(Q) — 시간 단위 숫자로 기록(양식 수식이 그대로 빼므로).
        hours = str(ex.get("hours", "") or "").strip()
        q_val = 0.0
        if payoff == "O" and hours:
            q_val = _to_hours(hours)
            xml = _set_cell(xml, f"Q{r}", "num", _nf(q_val))
        else:
            xml = _set_cell(xml, f"Q{r}", "str", "")  # 비움 → 수식이 0으로 처리

        # 제외할 시간(N)·제외 사유(O) — 사용자가 입력하면 신청시간에서 차감.
        exclude = str(ex.get("exclude", "") or "").strip()
        n_val = _to_hours(exclude)
        if n_val > 0:
            xml = _set_cell(xml, f"N{r}", "num", _nf(n_val))
        else:
            xml = _set_cell(xml, f"N{r}", "str", "")
        reason = str(ex.get("exclude_reason", "") or "").strip()
        xml = _set_cell(xml, f"O{r}", "str", reason)

        note = str(ex.get("note", "") or "").strip()
        if note:
            xml = _set_cell(xml, f"R{r}", "str", note)

        # 수식 결과(근무시간 K·신청시간 S·지급시간 T)를 양식 수식과 똑같이 미리 계산해
        # 캐시값으로 넣는다. 제한된 보기에서도 0 대신 실제 값이 보인다.
        eps = 1e-9
        frac = (j_sec - i_sec) / DAY_SECONDS          # K(근무시간): MOD(J-I,1) = 퇴근-근무시작
        mod1 = frac - math.floor(frac)
        k_val = math.floor(mod1 * 24 * 2 + eps) / 2   # 근무시간(0.5h 단위)
        # S(신청시간): (M-L)*24 - N - Q = (조기출근이면 퇴근-근무시작, 아니면 승인초과) - 제외 - 대체휴무
        base = max(0.0, claim_sec / 3600.0 - n_val - q_val)
        s_val = math.floor(base * 2 + eps) / 2
        total_s += s_val
        xml = _set_formula_cache(xml, f"K{r}", _nf(k_val))
        xml = _set_formula_cache(xml, f"S{r}", _nf(s_val))   # 신청시간
        xml = _set_formula_cache(xml, f"T{r}", _nf(s_val))   # 지급시간(=신청시간)

    # 합계(S12 신청 / T12 지급) 캐시값
    xml = _set_formula_cache(xml, "S12", _nf(total_s))
    xml = _set_formula_cache(xml, "T12", _nf(total_s))

    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                data = xml.encode("utf-8")
            elif item.filename == "xl/workbook.xml":
                # 열 때 신청시간·지급시간 등 모든 수식을 강제 재계산
                data = _force_full_recalc(data.decode("utf-8")).encode("utf-8")
            zi = zipfile.ZipInfo(item.filename, date_time=item.date_time)
            zi.compress_type = item.compress_type
            zi.external_attr = item.external_attr
            zi.internal_attr = item.internal_attr
            zi.create_system = item.create_system
            zout.writestr(zi, data)
    zin.close()
    out.seek(0)
    return out, len(records)


if __name__ == "__main__":
    import sys
    att = sys.argv[1] if len(sys.argv) > 1 else "남소희_월간근태현황_202605.xlsx"
    tpl = sys.argv[2] if len(sys.argv) > 2 else "초과근무(수당)신청서_양식.xlsx"
    name, year, month, recs, unapproved = parse_attendance(att)
    print(f"이름={name} 연={year} 월={month} 대상일수={len(recs)} 미승인합계(시간)={unapproved}")
    for rc in recs:
        print(rc)
    buf, n = fill_overtime(tpl, att)
    with open("test_overtime.xlsx", "wb") as f:
        f.write(buf.read())
    print(f"채움 완료: {n}건 -> test_overtime.xlsx")
